"""Built-in MCP server: skill-tools."""

import asyncio
import base64
import html
import json
import locale
import os
import signal
import shlex
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent

from agentclaw.mcp.file_lock import file_write_lock
from agentclaw.mcp.download_limits import (
    normalize_download_ttl,
    validate_download_file_size,
)
from agentclaw.platform_compat import run_subprocess_in_executor
from agentclaw.database.manager import DatabaseManager, RedisConfig
from agentclaw.logger.config import get_logger
from agentclaw.memory import (
    MEMORY_CHAR_LIMIT,
    MEMORY_COMPRESS_TARGET,
    get_workflow_memory_path,
    read_workflow_memory,
    update_workflow_memory,
    write_workflow_memory,
)

logger = get_logger(__name__)


def _add_line_numbers(content: str, start_line: int = 1) -> str:
    """给文本内容添加行号"""
    lines = content.split('\n')
    width = len(str(start_line + len(lines) - 1))
    numbered = []
    for i, line in enumerate(lines):
        line_no = start_line + i
        numbered.append(f"{line_no:>{width}}| {line}")
    return '\n'.join(numbered)


def _decode_process_output(data: Any) -> str:
    """Decode subprocess output without assuming every platform emits UTF-8."""
    if data is None:
        return ""
    if isinstance(data, str):
        return data
    if not isinstance(data, (bytes, bytearray)):
        return str(data)

    raw = bytes(data)
    if not raw:
        return ""

    encodings: list[str] = []

    def add_encoding(encoding: str | None) -> None:
        if not encoding:
            return
        normalized = encoding.strip()
        if normalized and normalized.lower() not in {item.lower() for item in encodings}:
            encodings.append(normalized)

    add_encoding("utf-8-sig")
    add_encoding("utf-8")
    add_encoding(locale.getpreferredencoding(False))

    if sys.platform == "win32":
        add_encoding("oem")
        add_encoding("mbcs")
        add_encoding("cp65001")
        add_encoding("cp936")
        add_encoding("gbk")
        add_encoding("cp950")
        add_encoding("big5")
        add_encoding("cp932")
        add_encoding("shift_jis")
        add_encoding("cp949")

    for encoding in encodings:
        try:
            return raw.decode(encoding)
        except (LookupError, UnicodeDecodeError):
            continue

    return raw.decode("utf-8", errors="replace")


def _strip_wrapping_quotes(value: str) -> str:
    if len(value) >= 2 and value[0] == value[-1] and value[0] in ("'", '"'):
        return value[1:-1]
    return value


def _windows_executable_basename(executable: str) -> str:
    return executable.replace("\\", "/").rsplit("/", 1)[-1].lower()


def _build_windows_powershell_encoded_command(command: str) -> Optional[list[str]]:
    """Convert explicit PowerShell -Command invocations to UTF-16LE EncodedCommand.

    Running `powershell -Command "...non-ascii..."` through `cmd.exe /c` can
    corrupt CJK literals on Windows code pages before the HTTP/API layer sees
    them. `-EncodedCommand` keeps the script payload in PowerShell's documented
    UTF-16LE format and bypasses the cmd.exe command-line code page boundary.
    """
    try:
        parts = shlex.split(command, posix=False)
    except ValueError:
        return None
    if not parts:
        return None

    executable = _strip_wrapping_quotes(parts[0])
    if _windows_executable_basename(executable) not in {"powershell", "powershell.exe", "pwsh", "pwsh.exe"}:
        return None

    command_index = -1
    for index, part in enumerate(parts[1:], start=1):
        normalized = _strip_wrapping_quotes(part).lower()
        if normalized in {"-encodedcommand", "-enc", "-e"}:
            return None
        if normalized in {"-command", "-c"}:
            command_index = index
            break

    if command_index < 0 or command_index + 1 >= len(parts):
        return None

    script = _strip_wrapping_quotes(" ".join(parts[command_index + 1:]).strip())
    if not script:
        return None

    utf8_prelude = (
        "$__agentclawUtf8 = New-Object System.Text.UTF8Encoding $false; "
        "[Console]::OutputEncoding = $__agentclawUtf8; "
        "$OutputEncoding = $__agentclawUtf8; "
    )
    encoded = base64.b64encode((utf8_prelude + script).encode("utf-16le")).decode("ascii")
    prefix_args = [_strip_wrapping_quotes(part) for part in parts[1:command_index]]
    return [executable, *prefix_args, "-EncodedCommand", encoded]


def _format_exception(exc: BaseException) -> str:
    """Format exceptions so empty messages still carry useful diagnostics."""
    message = str(exc).strip()
    exc_name = type(exc).__name__
    return f"{exc_name}: {message}" if message else exc_name


def _read_timeout_arg(args: dict, default: float = 120.0) -> float:
    """Read a positive timeout from tool arguments."""
    raw = args.get("timeout")
    if raw is None:
        return default
    try:
        timeout = float(raw)
    except (TypeError, ValueError):
        return default
    return timeout if timeout > 0 else default


def _format_seconds(seconds: float) -> str:
    return f"{seconds:g}"


async def _terminate_async_process_tree(proc: asyncio.subprocess.Process) -> None:
    """Terminate an asyncio subprocess and its child process group when possible."""
    if proc.returncode is not None:
        return
    if sys.platform != "win32":
        try:
            os.killpg(proc.pid, signal.SIGKILL)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass
    else:
        try:
            proc.kill()
        except Exception:
            pass
    try:
        await asyncio.wait_for(proc.communicate(), timeout=2.0)
    except Exception:
        pass


class SkillToolsServer:
    """
    Skill Tools MCP Server
    
    Provides code execution tools with skill environment support:
    - python: Execute Python code/script with skill's venv
    - shell: Execute shell commands
    - read_file: Read file content
    - write_file: Write file content
    - list_files: List directory contents
    
    Usage:
        python -m agentclaw.mcp.builtin_servers skill-tools \\
            --skills-dir ./skills --working-dir .
    """
    
    def __init__(
        self,
        skills_dir: Optional[str] = None,
        working_dir: Optional[str] = None,
        models_config: Optional[str] = None,
        project_dir: Optional[str] = None,
    ):
        self.skills_dir = Path(skills_dir) if skills_dir else None
        self.working_dir = Path(working_dir) if working_dir else Path.cwd()
        self.models_config = Path(models_config) if models_config else None
        self.project_dir = Path(project_dir) if project_dir else self.working_dir
        self._skill_manager = None
        self._skill_manager_init_attempted = False
        self._vl_client = None  # lazy init
        self._default_llm_manager = None
        self._server = Server("skill-tools")
        self._write_file_max_bytes = self._read_positive_int_env(
            "SKILL_TOOLS_WRITE_FILE_MAX_BYTES",
            200000,
        )
        self._document_read_timeout = self._read_positive_float_env(
            "SKILL_TOOLS_DOCUMENT_READ_TIMEOUT",
            30.0,
        )
        self._setup_handlers()

    @staticmethod
    def _read_positive_int_env(name: str, default: int) -> int:
        raw = os.getenv(name)
        if raw is None:
            return default
        try:
            parsed = int(raw)
        except Exception:
            return default
        return parsed if parsed > 0 else default

    @staticmethod
    def _read_positive_float_env(name: str, default: float) -> float:
        raw = os.getenv(name)
        if raw is None:
            return default
        try:
            parsed = float(raw)
        except Exception:
            return default
        return parsed if parsed > 0 else default

    def _resolve_server_base_url(self, *, internal: bool = False) -> str:
        """Resolve the server base URL from relay config or environment."""
        project_dir = Path(os.getenv("AGENTCLAW_PROJECT_DIR") or self.project_dir).expanduser().resolve()
        relay_cfg = project_dir / ".agentclaw" / "relay.json"
        if relay_cfg.exists():
            try:
                import json as _j
                cfg = _j.loads(relay_cfg.read_text(encoding="utf-8"))
                if cfg.get("project_dir") and Path(cfg["project_dir"]).expanduser().resolve() != project_dir:
                    raise ValueError("relay config belongs to a different project")
                key = "internal_url" if internal else "url"
                url = cfg.get(key) or cfg.get("url", "")
                if url:
                    return url.rstrip("/")
            except Exception:
                pass
        if os.getenv("AGENTCLAW_ALLOW_GLOBAL_RELAY_CONFIG", "").lower() in {"1", "true", "yes"}:
            relay_cfg = Path(tempfile.gettempdir()) / ".agentclaw_relay.json"
            if relay_cfg.exists():
                try:
                    import json as _j
                    cfg = _j.loads(relay_cfg.read_text(encoding="utf-8"))
                    key = "internal_url" if internal else "url"
                    url = cfg.get(key) or cfg.get("url", "")
                    if url:
                        return url.rstrip("/")
                except Exception:
                    pass
        if internal:
            env_internal_url = os.getenv("AGENTCLAW_INTERNAL_URL", "").strip()
            if env_internal_url:
                return env_internal_url.rstrip("/")
        env_url = os.getenv("AGENTCLAW_URL", "").strip()
        if env_url:
            return env_url.rstrip("/")
        return "http://127.0.0.1:8000"

    def _substitute_skill_runtime_variables(self, content: str, skill_name: str) -> str:
        """Inject runtime URLs into skill docs.

        Agent-facing platform API skills use the local-only internal relay as
        their base URL; public/user-facing docs keep the normal server URL.
        """
        if "{BASE_URL}" not in content and "{INTERNAL_BASE_URL}" not in content:
            return content

        internal_skill_names = {"agentclaw_api", "agent_creator"}
        base_url = self._resolve_server_base_url(
            internal=skill_name in internal_skill_names,
        )
        internal_base_url = self._resolve_server_base_url(internal=True)

        return (
            content
            .replace("{INTERNAL_BASE_URL}", internal_base_url)
            .replace("{BASE_URL}", base_url)
        )

    async def _init_skill_manager(self):
        """Initialize SkillManager if skills_dir is provided"""
        if self._skill_manager is not None:
            return
        if self._skill_manager_init_attempted:
            return

        if not self.skills_dir or not self.skills_dir.exists():
            return

        self._skill_manager_init_attempted = True
        try:
            # Run skill loading in thread pool to avoid blocking event loop
            loop = asyncio.get_running_loop()
            await loop.run_in_executor(None, self._load_skills_sync)
            logger.info(f"[skill-tools] Loaded {len(self._skill_manager)} skills")
        except Exception as e:
            logger.warning(f"[skill-tools] Failed to init SkillManager: {e}")

    def _load_skills_sync(self):
        """Synchronous skill loading (runs in thread pool)"""
        from agentclaw.skills.manager import SkillManager
        self._skill_manager = SkillManager(str(self.skills_dir), auto_init_env=True)
        self._skill_manager.load_all()

    def _get_builtin_skill_manager(self):
        """Get builtin skill manager (lazy init)"""
        if not hasattr(self, "_builtin_skill_manager"):
            try:
                from agentclaw.skills import get_builtin_skill_manager
                self._builtin_skill_manager = get_builtin_skill_manager(auto_init=True)
            except Exception:
                self._builtin_skill_manager = None
        return self._builtin_skill_manager

    def _get_skill_binding(self, skill_name: str):
        """Resolve a skill by name from project skills or builtin skills."""
        if not skill_name:
            return None, None

        if self._skill_manager:
            skill_obj = self._skill_manager.get(skill_name)
            if skill_obj:
                return skill_obj, self._skill_manager

        builtin_mgr = self._get_builtin_skill_manager()
        if builtin_mgr:
            skill_obj = builtin_mgr.get(skill_name)
            if skill_obj:
                return skill_obj, builtin_mgr

        return None, None

    async def _run_exec_process(
        self,
        cmd: list[str],
        *,
        cwd: str,
        env: dict[str, str],
        timeout: float,
    ) -> subprocess.CompletedProcess | tuple[Any, Any, int]:
        if sys.platform == "win32":
            return await run_subprocess_in_executor(
                cmd,
                cwd=cwd,
                env=env,
                timeout=timeout,
            )
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdin=asyncio.subprocess.DEVNULL,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=cwd,
            env=env,
            start_new_session=True,
        )
        try:
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        except asyncio.TimeoutError:
            await _terminate_async_process_tree(proc)
            raise
        return stdout, stderr, proc.returncode

    async def _run_shell_process(
        self,
        command: str,
        *,
        cwd: str,
        env: dict[str, str],
        timeout: float,
        executable: Optional[str] = None,
    ) -> subprocess.CompletedProcess | tuple[Any, Any, int]:
        if sys.platform == "win32":
            return await run_subprocess_in_executor(
                command,
                cwd=cwd,
                env=env,
                timeout=timeout,
                shell=True,
                executable=executable,
            )
        proc = await asyncio.create_subprocess_shell(
            command,
            stdin=asyncio.subprocess.DEVNULL,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            cwd=cwd,
            env=env,
            executable=executable,
            start_new_session=True,
        )
        try:
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        except asyncio.TimeoutError:
            await _terminate_async_process_tree(proc)
            raise
        return stdout, stderr, proc.returncode
    
    def _setup_handlers(self):
        """Setup MCP server handlers"""
        
        @self._server.list_tools()
        async def list_tools() -> List[Tool]:
            return [
                Tool(
                    name="python",
                    description=(
                        "Execute Python code or script file. MUST provide either 'code' OR 'file'. "
                        "For code payloads, prefer real multi-line content (avoid literal \\\\n strings). "
                        "Prefer multi-line try/except blocks for reliability. "
                        "Note: code mode runs as a temporary script; if importing project modules (e.g. workflows.*), prepend sys.path with os.getcwd() or verify via shell python -c in project cwd. "
                        "When calling Python through shell, match the runtime platform: Windows usually uses `python` or `py -3`; Linux/macOS usually use `python3`."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "code": {
                                "type": "string",
                                "description": "Python code to execute (required if 'file' not provided)"
                            },
                            "file": {
                                "type": "string",
                                "description": "Path to .py file to execute (required if 'code' not provided)"
                            },
                            "skill_name": {
                                "type": "string",
                                "description": "Skill name for skill-relative script execution and skill environment resolution (optional)"
                            },
                            "skill": {
                                "type": "string",
                                "description": "Deprecated alias of skill_name, kept for compatibility"
                            },
                            "args": {
                                "description": "Command line arguments passed to the script. Pass as JSON array of strings, e.g. [\"arg1\", \"--flag\", \"value\"]. Use [] if no arguments needed. Check the script's usage or skill docs for required arguments."
                            },
                            "timeout": {
                                "type": "number",
                                "description": "Optional timeout in seconds. Defaults to 120 if omitted. Use a small value for risky code that may install packages or spawn subprocesses."
                            }
                        },
                        "required": ["args"]
                    }
                ),
                Tool(
                    name="javascript",
                    description="Execute JavaScript/Node.js code or script file. MUST provide either 'code' OR 'file' parameter.",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "code": {
                                "type": "string",
                                "description": "JavaScript code to execute (required if 'file' not provided)"
                            },
                            "file": {
                                "type": "string",
                                "description": "Path to .js file to execute (required if 'code' not provided)"
                            },
                            "skill_name": {
                                "type": "string",
                                "description": "Skill name for skill-relative script execution and skill dependency resolution (optional)"
                            },
                            "skill": {
                                "type": "string",
                                "description": "Deprecated alias of skill_name, kept for compatibility"
                            },
                            "args": {
                                "description": "Command line arguments passed to the script. Pass as JSON array of strings, e.g. [\"arg1\", \"--flag\", \"value\"]. Use [] if no arguments needed. Check the script's usage or skill docs for required arguments."
                            },
                            "timeout": {
                                "type": "number",
                                "description": "Optional timeout in seconds. Defaults to 120 if omitted."
                            }
                        },
                        "required": ["args"]
                    }
                ),
                Tool(
                    name="shell",
                    description="Execute shell command in the working directory. Match command syntax to the current runtime platform. On Windows this runs via cmd.exe; on Linux/macOS it uses the default POSIX shell. All paths are relative to working directory.",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "command": {
                                "type": "string",
                                "description": "Shell command to execute (use relative paths, no need to cd)"
                            },
                            "timeout": {
                                "type": "number",
                                "description": "Optional timeout in seconds. Defaults to 120 if omitted."
                            }
                        },
                        "required": ["command"]
                    }
                ),
                Tool(
                    name="read_file",
                    description="Preferred tool for reading existing local files relative to skill-tools working_dir. "
                                "Use this for file inspection before choosing python or shell for computation or command execution. "
                                "Supports plain text, PDF, DOCX, PPTX, XLSX (converted to markdown), "
                                "and images (analyzed via vision model if configured). "
                                "Pass path for one file or paths for multiple files in one call. "
                                "Document conversion runs in an isolated subprocess and times out via "
                                "SKILL_TOOLS_DOCUMENT_READ_TIMEOUT (default 30s). "
                                "Output includes line numbers for reference. "
                                "If content is truncated, use line_start to read remaining lines. "
                                "Note: working_dir can differ from coding-tools project_dir. "
                                "Path resolution tries working_dir first, then ancestor directories. "
                                "Pass skill_name to resolve path relative to a skill's root directory.",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "path": {
                                "type": "string",
                                "description": "File path relative to working directory"
                            },
                            "paths": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": "Multiple file paths relative to working directory. Use when comparing or summarizing several files in one call."
                            },
                            "prompt": {
                                "type": "string",
                                "description": "Vision prompt for image files ONLY (e.g. PNG/JPG). Ignored for non-image files. Do NOT pass for text/code files."
                            },
                            "skill_name": {
                                "type": "string",
                                "description": "Skill name to resolve path relative to skill's root directory (optional)"
                            },
                            "line_start": {
                                "type": "integer",
                                "description": "Start reading from this line number (1-based). Use when previous read was truncated to continue from the truncation point."
                            }
                        }
                    }
                ),
                Tool(
                    name="update_memory",
                    description=(
                        "Update the current workflow memory.md using regex replacement. "
                        "Supports adding, deleting, or modifying memory entries. "
                        "By default only the first regex match is replaced to avoid accidental duplication; "
                        "set replace_all=true only when you intentionally want every match updated. "
                        "Pass workflow_id explicitly, or let runtime inject the current workflow automatically."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "workflow_id": {"type": "string", "description": "Workflow ID (optional if runtime injects current workflow)"},
                            "pattern": {"type": "string", "description": "Regex pattern to match in memory.md"},
                            "replacement": {"type": "string", "description": "Replacement text"},
                            "dry_run": {"type": "boolean", "description": "Preview only", "default": False},
                            "replace_all": {"type": "boolean", "description": "Replace every regex match instead of only the first one", "default": False},
                        },
                        "required": ["pattern", "replacement"]
                    }
                ),
                Tool(
                    name="compress_memory",
                    description=(
                        "Compress the current workflow memory.md with the default model. "
                        f"Output must stay under {MEMORY_COMPRESS_TARGET} characters without losing key facts."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "workflow_id": {"type": "string", "description": "Workflow ID (optional if runtime injects current workflow)"},
                            "dry_run": {"type": "boolean", "description": "Preview only", "default": False},
                        },
                    }
                ),
                Tool(
                    name="write_file",
                    description=(
                        "Write content to a file relative to skill-tools working_dir. "
                        "Content is written LITERALLY as-is — no escape sequence processing "
                        "(\\n in content becomes literal backslash-n, not a newline). "
                        "To write actual newlines, use real multi-line content. "
                        "path must be a relative file path, content must be source/text payload. "
                        "Do not swap path/content. "
                        "For existing files, overwrite requires explicit overwrite_existing=true. "
                        "For local edits on existing files, prefer coding-tools replace_in_file/update_code. "
                        "If using coding-tools after write_file, re-check path base because coding-tools project_dir may differ. "
                        "Result contract: success returns [OK] with written path; parameter/path validation failures return [Error] with structured root_cause_hint."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "path": {
                                "type": "string",
                                "description": (
                                    "File path relative to working directory "
                                    "(e.g., 'workflows/nl2sql.py'). Must not be absolute path, "
                                    "must not contain newlines/code snippets."
                                )
                            },
                            "content": {
                                "type": "string",
                                "description": "File content to write (written literally, no escape processing)"
                            },
                            "overwrite_existing": {
                                "type": "boolean",
                                "description": (
                                    "Explicitly allow overwriting an existing file. "
                                    "Default false to encourage local patch tools for existing files."
                                ),
                                "default": False
                            }
                        },
                        "required": ["path", "content"]
                    }
                ),
                Tool(
                    name="write_code",
                    description=(
                        "Write code to a file with automatic sanitization. "
                        "Preferred over write_file for source code (.py, .js, .ts, .json, etc). "
                        "Auto-fixes HTML entities from model output (&quot; -> \", &#x27; -> ', &gt; -> >). "
                        "Runs syntax check before writing — rejects invalid code without touching the file. "
                        "path must be relative. content is the full source code. "
                        "Result contract: [OK] on success; [SYNTAX_ERROR] with context on pre-write check failure; [Error] on validation failure."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "path": {
                                "type": "string",
                                "description": (
                                    "File path relative to working directory "
                                    "(e.g., 'workflows/nl2sql.py'). Must not be absolute."
                                )
                            },
                            "content": {
                                "type": "string",
                                "description": "Source code content to write"
                            },
                            "overwrite_existing": {
                                "type": "boolean",
                                "description": "Allow overwriting an existing file (default false)",
                                "default": False
                            }
                        },
                        "required": ["path", "content"]
                    }
                ),
                Tool(
                    name="list_files",
                    description=(
                        "List directory contents under skill-tools working_dir. "
                        "Path resolution tries working_dir first, then ancestor directories. "
                        "If path resolves to a file, returns file info with hint to use read_file. "
                        "Pass skill_name to list files relative to a skill's root directory."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "path": {
                                "type": "string",
                                "description": "Directory path (default: current directory)",
                                "default": "."
                            },
                            "skill_name": {
                                "type": "string",
                                "description": "Skill name to resolve path relative to skill's root directory (optional)"
                            }
                        }
                    }
                ),
                Tool(
                    name="read_skill_file",
                    description=(
                        "Read a skill's documentation file by skill name. "
                        "Searches in both project skills and builtin skills directories. "
                        "Use this to read SKILL.md or reference files from skills."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "skill_name": {
                                "type": "string",
                                "description": "Name of the skill (e.g., 'agent_creator', 'coding_skill', 'gemini')"
                            },
                            "file_name": {
                                "type": "string",
                                "description": "File name to read (default: 'SKILL.md')",
                                "default": "SKILL.md"
                            }
                        },
                        "required": ["skill_name"]
                    }
                ),
                Tool(
                    name="execute_sudo_command",
                    description=(
                        "Execute a command with sudo privileges. "
                        "When Harness tool confirmation is enabled, this tool will request user authorization and sudo password before execution. "
                        "Use this for commands that require root access like docker, systemctl, apt, etc."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "command": {
                                "type": "string",
                                "description": "The command to execute with sudo (without 'sudo' prefix), e.g. 'docker ps -a'"
                            },
                            "timeout": {
                                "type": "number",
                                "description": "Optional timeout in seconds. Defaults to 60 if omitted."
                            }
                        },
                        "required": ["command"]
                    }
                ),
                Tool(
                    name="create_download_url",
                    description=(
                        "Generate a temporary download URL for a local file. "
                        "Users can download/view the file (images, PDFs, documents) via this URL. "
                        "Link expires after 1 hour by default. Requires Redis."
                    ),
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "path": {
                                "type": "string",
                                "description": "Local file path (absolute or relative to working directory)"
                            },
                            "filename": {
                                "type": "string",
                                "description": "Display filename for download (optional, defaults to original filename)"
                            },
                            "ttl": {
                                "type": "integer",
                                "description": "Link validity period in seconds (default 3600)",
                                "default": 3600
                            }
                        },
                        "required": ["path"]
                    }
                ),
            ]
        
        @self._server.call_tool()
        async def call_tool(name: str, arguments: dict) -> List[TextContent]:
            try:
                # 在结果前添加工作目录信息（帮助 LLM 理解环境）
                cwd_info = f"[Working Directory: {self.working_dir}]\n"
                
                if name == "python":
                    result = await self._execute_python(arguments)
                elif name == "javascript":
                    result = await self._execute_javascript(arguments)
                elif name == "shell":
                    result = await self._execute_shell(arguments)
                elif name == "read_file":
                    result = await self._read_file(arguments)
                elif name == "update_memory":
                    result = await self._update_memory(arguments)
                elif name == "compress_memory":
                    result = await self._compress_memory(arguments)
                elif name == "write_file":
                    result = await self._write_file(arguments)
                elif name == "write_code":
                    result = await self._write_code(arguments)
                elif name == "list_files":
                    result = await self._list_files(arguments)
                elif name == "read_skill_file":
                    result = await self._read_skill_file(arguments)
                elif name == "execute_sudo_command":
                    result = await self._execute_sudo_command(arguments)
                elif name == "create_download_url":
                    result = await self._create_download_url(arguments)
                else:
                    result = f"Unknown tool: {name}"
                
                # 对于文件操作，添加工作目录提示
                if name in ("python", "javascript", "shell", "write_file", "write_code", "execute_sudo_command"):
                    result = cwd_info + result
                
                return [TextContent(type="text", text=str(result))]
            except Exception as e:
                return [TextContent(type="text", text=f"Error: {e}")]
    
    async def _execute_python(self, args: dict) -> str:
        """Execute Python code or script"""
        code = args.get("code")
        file = args.get("file")
        timeout = _read_timeout_arg(args)
        skill_name = (args.get("skill_name") or args.get("skill") or "").strip()
        # 兼容各种格式: list, string (JSON array), string (plain), number, None
        raw_args = args.get("args", [])
        if raw_args is None:
            cmd_args = []
        elif isinstance(raw_args, list):
            cmd_args = [str(a) for a in raw_args]
        elif isinstance(raw_args, str):
            # LLM 可能将数组序列化为字符串，尝试 JSON 解析
            try:
                parsed = json.loads(raw_args)
                if isinstance(parsed, list):
                    cmd_args = [str(a) for a in parsed]
                else:
                    cmd_args = [raw_args] if raw_args.strip() else []
            except (json.JSONDecodeError, ValueError):
                cmd_args = [raw_args] if raw_args.strip() else []
        else:
            cmd_args = [str(raw_args)]

        if not code and not file:
            return "[Error] Must provide either 'code' or 'file' parameter"

        if code and file:
            return "[Error] Cannot use both 'code' and 'file', choose one"

        # Determine Python interpreter and environment
        python_cmd = sys.executable
        env_vars = os.environ.copy()
        env_vars["PYTHONIOENCODING"] = "utf-8"
        env_vars.setdefault("PYTHONUTF8", "1")
        skill_path: Optional[Path] = None

        # 确保 venv 的 bin/Scripts 在 PATH 最前面，使 pip/python 指向 venv
        venv_bin = Path(sys.executable).parent
        current_path = env_vars.get("PATH", "")
        if str(venv_bin) not in current_path:
            env_vars["PATH"] = str(venv_bin) + os.pathsep + current_path
        # 设置 VIRTUAL_ENV 使 pip 知道自己在 venv 中
        venv_root = venv_bin.parent
        if (venv_root / "pyvenv.cfg").exists():
            env_vars["VIRTUAL_ENV"] = str(venv_root)
        
        if skill_name:
            # 先从项目 skill manager 查找，再从 builtin skill manager 查找
            skill_obj, active_manager = self._get_skill_binding(skill_name)

            if skill_obj:
                skill_path = skill_obj.path

            # 尝试初始化 skill 虚拟环境（无 requirements.txt 时跳过，使用系统 Python）
            if active_manager:
                env = active_manager.get_env(skill_name)
                if env and env.is_initialized:
                    python_cmd = env.python
                elif skill_obj and skill_obj.requirements_file:
                    # 有 requirements.txt 但 env 未初始化，尝试初始化
                    if active_manager.init_env(skill_name):
                        env = active_manager.get_env(skill_name)
                        if env:
                            python_cmd = env.python
                    else:
                        return f"[Error] Failed to initialize skill '{skill_name}' environment"
                # 无 requirements.txt 的 skill，直接使用系统 Python

            if not skill_path or not skill_path.exists():
                return f"[Error] Skill not found: {skill_name}"

            # Set PYTHONPATH
            pythonpath_parts = [str(skill_path)]
            for subdir in ["scripts", "core"]:
                sub_path = skill_path / subdir
                if sub_path.exists():
                    pythonpath_parts.append(str(sub_path))
            if env_vars.get("PYTHONPATH"):
                pythonpath_parts.append(env_vars["PYTHONPATH"])
            env_vars["PYTHONPATH"] = os.pathsep.join(pythonpath_parts)
        
        # Prepare script
        temp_script = None
        if code:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8') as f:
                f.write(code)
                script_path = f.name
                temp_script = script_path
        else:
            file_path = Path(file)
            script_path_obj: Optional[Path] = None

            if skill_path:
                # 有 skill_name：skill 目录优先 → fallback working_dir → 绝对路径兼容
                if not file_path.is_absolute():
                    candidate = (skill_path / file_path).resolve()
                    if candidate.exists() and candidate.is_file():
                        script_path_obj = candidate
                    else:
                        # fallback 到 working_dir
                        fallback = (self.working_dir / file_path).resolve()
                        if fallback.exists() and fallback.is_file():
                            script_path_obj = fallback
                else:
                    # 绝对路径直接用
                    script_path_obj = file_path
            else:
                # 无 skill_name：working_dir 相对路径 / 绝对路径
                if file_path.is_absolute():
                    script_path_obj = file_path
                else:
                    script_path_obj = (self.working_dir / file_path).resolve()

            if not script_path_obj or not script_path_obj.exists() or not script_path_obj.is_file():
                return f"[Error] File not found: {script_path_obj if script_path_obj else file}"

            script_path = str(script_path_obj)
        
        try:
            cmd = [python_cmd, script_path] + cmd_args
            exec_result = await self._run_exec_process(
                cmd,
                cwd=str(self.working_dir),
                env=env_vars,
                timeout=timeout,
            )
            stdout, stderr, returncode = (
                (exec_result.stdout, exec_result.stderr, exec_result.returncode)
                if sys.platform == "win32"
                else exec_result
            )

            stdout_text = _decode_process_output(stdout)
            stderr_text = _decode_process_output(stderr)
            if returncode != 0:
                # 检测 argparse 帮助/用法输出（stdout 或 stderr 以 "usage:" 开头）
                # 这不是真正的错误，不应触发 LLM 重试
                combined = (stdout_text + stderr_text).strip().lower()
                is_argparse_usage = (
                    returncode == 2
                    or combined.startswith("usage:")
                    or "the following arguments are required" in combined
                )
                if is_argparse_usage:
                    output = stdout_text + stderr_text
                    return (
                        f"[ACTION REQUIRED] Script needs command-line arguments. You MUST re-call with the 'args' parameter.\n"
                        f"{output}\n"
                        f"[EXAMPLE] {{"
                        f'"file": "{file}", '
                        f'"skill_name": "{skill_name}", '
                        f'"args": ["<positional_arg>", "--flag", "value"]'
                        f"}}\n"
                        f"Read the usage above to determine the correct args, then re-call immediately."
                    )

                parts = [f"[ERROR] Python execution failed (exit code {returncode})"]
                if stdout_text.strip():
                    parts.append(f"[stdout]\n{stdout_text}")
                if stderr_text.strip():
                    parts.append(f"[stderr]\n{stderr_text}")
                if (
                    "ModuleNotFoundError: No module named 'workflows'" in stderr_text
                    or 'ModuleNotFoundError: No module named "workflows"' in stderr_text
                ):
                    parts.append(
                        "[hint] Project module import failed in temp-script context. "
                        "Try adding:\n"
                        "import os, sys\n"
                        "sys.path.insert(0, os.getcwd())\n"
                        "before importing workflows.* (or verify with the platform's Python command, "
                        "for example Linux/macOS: python3 -c 'import workflows.xxx', Windows: python -c 'import workflows.xxx')."
                    )
                return "\n".join(parts)

            result = stdout_text
            if stderr_text:
                result += f"\n[stderr]\n{stderr_text}"
            return result if result else "(no output)"
        except asyncio.TimeoutError:
            return (
                f"Execution timeout ({_format_seconds(timeout)}s)\n"
                "[hint] The Python tool terminated the process tree. If the script installs packages with pip or runs long child processes, "
                "prefer the dedicated install_package tool first, or re-run with a larger timeout only when necessary."
            )
        finally:
            if temp_script:
                os.unlink(temp_script)
    
    async def _execute_javascript(self, args: dict) -> str:
        """Execute JavaScript/Node.js code or script"""
        code = args.get("code")
        file = args.get("file")
        timeout = _read_timeout_arg(args)
        skill_name = (args.get("skill_name") or args.get("skill") or "").strip()
        # 兼容各种格式: list, string (JSON array), string (plain), number, None
        raw_args = args.get("args", [])
        if raw_args is None:
            cmd_args = []
        elif isinstance(raw_args, list):
            cmd_args = [str(a) for a in raw_args]
        elif isinstance(raw_args, str):
            try:
                parsed = json.loads(raw_args)
                if isinstance(parsed, list):
                    cmd_args = [str(a) for a in parsed]
                else:
                    cmd_args = [raw_args] if raw_args.strip() else []
            except (json.JSONDecodeError, ValueError):
                cmd_args = [raw_args] if raw_args.strip() else []
        else:
            cmd_args = [str(raw_args)]

        if not code and not file:
            return "[Error] Must provide either 'code' or 'file' parameter"

        if code and file:
            return "[Error] Cannot use both 'code' and 'file', choose one"

        # Find node executable
        node_cmd = "node"
        env_vars = os.environ.copy()
        skill_path: Optional[Path] = None
        
        # Build NODE_PATH: working_dir/node_modules + skill/node_modules
        node_path_parts = []
        
        # Always include working directory's node_modules
        working_node_modules = self.working_dir / "node_modules"
        if working_node_modules.exists():
            node_path_parts.append(str(working_node_modules))
        
        # Check if skill has node_modules
        if skill_name:
            skill_obj, _ = self._get_skill_binding(skill_name)
            if skill_obj:
                skill_path = skill_obj.path

            if not skill_path or not skill_path.exists():
                return f"[Error] Skill not found: {skill_name}"

            # Add skill's node_modules to NODE_PATH
            if skill_path:
                skill_node_modules = skill_path / "node_modules"
                if skill_node_modules.exists():
                    node_path_parts.append(str(skill_node_modules))

        # Set NODE_PATH
        if node_path_parts:
            existing_node_path = env_vars.get("NODE_PATH", "")
            if existing_node_path:
                node_path_parts.append(existing_node_path)
            env_vars["NODE_PATH"] = os.pathsep.join(node_path_parts)

        # Prepare script - create temp file in working directory so require() works
        temp_script = None
        if code:
            # Create temp file in working directory (not /tmp) for proper module resolution
            temp_file = self.working_dir / f"_temp_{os.getpid()}.js"
            temp_file.write_text(code, encoding='utf-8')
            script_path = str(temp_file)
            temp_script = script_path
        else:
            file_path = Path(file)
            script_path_obj = None

            if skill_path:
                # 有 skill_name：skill 目录优先 → fallback working_dir → 绝对路径兼容
                if not file_path.is_absolute():
                    candidate = (skill_path / file_path).resolve()
                    if candidate.exists() and candidate.is_file():
                        script_path_obj = candidate
                    else:
                        fallback = (self.working_dir / file_path).resolve()
                        if fallback.exists() and fallback.is_file():
                            script_path_obj = fallback
                else:
                    script_path_obj = file_path
            else:
                if file_path.is_absolute():
                    script_path_obj = file_path
                else:
                    script_path_obj = (self.working_dir / file_path).resolve()

            if not script_path_obj or not Path(script_path_obj).exists():
                return f"[Error] File not found: {file}"

            script_path = str(script_path_obj)
        
        try:
            cmd = [node_cmd, script_path] + cmd_args
            exec_result = await self._run_exec_process(
                cmd,
                cwd=str(self.working_dir),
                env=env_vars,
                timeout=timeout,
            )
            stdout, stderr, returncode = (
                (exec_result.stdout, exec_result.stderr, exec_result.returncode)
                if sys.platform == "win32"
                else exec_result
            )

            stdout_text = _decode_process_output(stdout)
            stderr_text = _decode_process_output(stderr)
            if returncode != 0:
                parts = [f"[ERROR] JavaScript execution failed (exit code {returncode})"]
                if stdout_text.strip():
                    parts.append(f"[stdout]\n{stdout_text}")
                if stderr_text.strip():
                    parts.append(f"[stderr]\n{stderr_text}")
                return "\n".join(parts)

            result = stdout_text
            if stderr_text:
                result += f"\n[stderr]\n{stderr_text}"

            return result if result else "(no output)"
        except asyncio.TimeoutError:
            return f"Execution timeout ({_format_seconds(timeout)}s)"
        except FileNotFoundError:
            return "[Error] Node.js not found. Please install Node.js to execute JavaScript code."
        finally:
            if temp_script:
                os.unlink(temp_script)
    
    async def _execute_shell(self, args: dict) -> str:
        """Execute shell command"""
        command = args.get("command", "")
        if not command:
            return "[Error] 'command' is required"

        # 解析超时参数
        timeout = _read_timeout_arg(args)

        # 构建 env：注入 venv PATH，确保 pip/python 指向正确
        env_vars = os.environ.copy()
        venv_bin = Path(sys.executable).parent
        current_path = env_vars.get("PATH", "")
        if str(venv_bin) not in current_path:
            env_vars["PATH"] = str(venv_bin) + os.pathsep + current_path
        venv_root = venv_bin.parent
        if (venv_root / "pyvenv.cfg").exists():
            env_vars["VIRTUAL_ENV"] = str(venv_root)

        powershell_exec_cmd = (
            _build_windows_powershell_encoded_command(command)
            if sys.platform == "win32"
            else None
        )

        # Windows: 使用 cmd.exe 或 powershell 而非默认的 /bin/sh
        shell_kwargs: dict = {}
        if sys.platform == "win32" and not powershell_exec_cmd:
            # 使用 cmd.exe /c 来执行命令
            shell_kwargs["executable"] = os.environ.get("COMSPEC", "cmd.exe")

        try:
            if powershell_exec_cmd:
                shell_result = await self._run_exec_process(
                    powershell_exec_cmd,
                    cwd=str(self.working_dir),
                    env=env_vars,
                    timeout=timeout,
                )
            else:
                shell_result = await self._run_shell_process(
                    command,
                    cwd=str(self.working_dir),
                    env=env_vars,
                    timeout=timeout,
                    executable=shell_kwargs.get("executable"),
                )
        except asyncio.TimeoutError:
            return f"Execution timeout ({_format_seconds(timeout)}s)"
        stdout, stderr, returncode = (
            (shell_result.stdout, shell_result.stderr, shell_result.returncode)
            if isinstance(shell_result, subprocess.CompletedProcess)
            else shell_result
        )

        stdout_text = _decode_process_output(stdout)
        stderr_text = _decode_process_output(stderr)
        if returncode != 0:
            parts = [f"[ERROR] Shell command failed (exit code {returncode})"]
            if stdout_text.strip():
                parts.append(f"[stdout]\n{stdout_text}")
            if stderr_text.strip():
                parts.append(f"[stderr]\n{stderr_text}")
            if (
                "python -c" in command
                and ("unterminated string literal" in stderr_text or "unexpected character after line continuation character" in stderr_text)
            ):
                parts.append(
                    "[hint] Detected likely quoting issue in inline Python. "
                    "Prefer python tool, or use shell with the platform's Python command. "
                    "Example Linux/macOS: python3 -c 'import workflows.contract_recommender'; "
                    "Windows: python -c 'import workflows.contract_recommender'"
                )
            return "\n".join(parts)

        result = stdout_text
        if stderr_text:
            result += f"\n[stderr]\n{stderr_text}"

        return result if result else "(no output)"
    
    # File extensions that markitdown can handle
    MARKITDOWN_EXTS = {'.pdf', '.docx', '.pptx', '.xlsx', '.xls', '.doc', '.ppt'}
    IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff', '.tif'}

    def _candidate_paths(self, path_str: str) -> list[Path]:
        """Build ordered path candidates (working_dir -> ancestor fallback)."""
        raw = Path(path_str)
        candidates: list[Path] = []

        def add_candidate(p: Path) -> None:
            try:
                resolved = p.resolve(strict=False)
            except Exception:
                resolved = p
            key = str(resolved)
            if all(str(existing) != key for existing in candidates):
                candidates.append(resolved)

        if raw.is_absolute():
            add_candidate(raw)
            return candidates

        add_candidate(self.working_dir / raw)
        for ancestor in self.working_dir.parents:
            add_candidate(ancestor / raw)

        return candidates

    def _resolve_existing_path(self, path_str: str) -> tuple[Optional[Path], list[Path]]:
        candidates = self._candidate_paths(path_str)
        for candidate in candidates:
            if candidate.exists():
                return candidate, candidates
        return None, candidates

    async def _read_file(self, args: dict) -> str:
        """Read file content with smart format handling"""
        path = args.get("path", "")
        raw_paths = args.get("paths")
        if raw_paths is None and isinstance(path, list):
            raw_paths = path
            path = ""
        if raw_paths is not None:
            if not isinstance(raw_paths, list):
                return "[Error] 'paths' must be a non-empty list of file paths"
            paths = [str(item).strip() for item in raw_paths if str(item).strip()]
            if not paths:
                raw_paths = None
            else:
                sections = []
                for item_path in paths:
                    child_args = dict(args)
                    child_args.pop("paths", None)
                    child_args["path"] = item_path
                    content = await self._read_file(child_args)
                    sections.append(f"===== {item_path} =====\n{content}")
                return "\n\n".join(sections)

        if not path:
            return "[Error] 'path' is required"

        skill_name = (args.get("skill_name") or "").strip()
        resolved_path = None
        candidates = []

        if skill_name:
            # skill 目录优先 → fallback working_dir
            skill_obj, _ = self._get_skill_binding(skill_name)
            if skill_obj and skill_obj.path and skill_obj.path.exists():
                skill_candidate = (skill_obj.path / path).resolve()
                if skill_candidate.exists():
                    resolved_path = skill_candidate
                candidates.append(skill_candidate)
            if not resolved_path:
                fallback, fallback_candidates = self._resolve_existing_path(path)
                candidates.extend(fallback_candidates)
                if fallback:
                    resolved_path = fallback
        else:
            resolved_path, candidates = self._resolve_existing_path(path)

        if not resolved_path:
            payload = {
                "ok": False,
                "error_class": "path_not_found",
                "requested_path": path,
                "working_dir": str(self.working_dir),
                "root_cause_hint": (
                    "path is resolved relative to skill-tools working_dir first, then ancestor directories"
                ),
                "suggested_fix": (
                    "Use a path relative to current working_dir, or an absolute existing path."
                ),
            }
            return "[Error] File not found\n" + json.dumps(payload, ensure_ascii=False, indent=2)
        if resolved_path.is_dir():
            return (
                f"[Error] Path is a directory, not a file: {path}\n"
                f"resolved_path={resolved_path}\n"
                "suggested_fix=Use list_files for directories, read_file for files."
            )

        file_path = resolved_path
        ext = file_path.suffix.lower()
        resolved_hint = ""
        direct_path = Path(path) if Path(path).is_absolute() else (self.working_dir / path)
        try:
            direct_resolved = direct_path.resolve(strict=False)
        except Exception:
            direct_resolved = direct_path
        if str(direct_resolved) != str(file_path):
            resolved_hint = f"[path_resolved] requested={path} resolved={file_path}\n"

        # Document files: use markitdown
        if ext in self.MARKITDOWN_EXTS:
            content = await self._read_document(file_path)
            return resolved_hint + content if resolved_hint else content

        # Image files: use VL model
        if ext in self.IMAGE_EXTS:
            prompt = args.get("prompt", "请详细描述这张图片的内容。")
            content = await self._read_image(file_path, prompt)
            return resolved_hint + content if resolved_hint else content

        # Plain text — add line numbers and support line_start offset
        line_start = args.get("line_start")
        try:
            content = file_path.read_text(encoding='utf-8')
            if line_start and isinstance(line_start, (int, float)) and int(line_start) > 1:
                lines = content.split('\n')
                start_idx = int(line_start) - 1
                if start_idx >= len(lines):
                    return f"{resolved_hint}[Info] line_start={int(line_start)} exceeds total lines ({len(lines)}). No content to show."
                content = '\n'.join(lines[start_idx:])
                numbered = _add_line_numbers(content, start_line=int(line_start))
            else:
                numbered = _add_line_numbers(content)
            return resolved_hint + numbered if resolved_hint else numbered
        except UnicodeDecodeError:
            content = file_path.read_bytes()
            prefix = resolved_hint if resolved_hint else ""
            return f"{prefix}(binary file, {len(content)} bytes)"
        except Exception as e:
            return f"Read failed: {e}"

    def _resolve_memory_workflow_id(self, args: dict) -> str:
        workflow_id = str(args.get("workflow_id", "")).strip()
        if not workflow_id:
            raise ValueError("workflow_id is required")
        return workflow_id

    def _get_default_llm_manager(self):
        if self._default_llm_manager is not None:
            return self._default_llm_manager
        from agentclaw.model.manager import LLMManager

        config_path = str(self.models_config) if self.models_config else "models.json"
        self._default_llm_manager = LLMManager(config_path=config_path)
        return self._default_llm_manager

    async def _update_memory(self, args: dict) -> str:
        try:
            workflow_id = self._resolve_memory_workflow_id(args)
        except ValueError as e:
            return f"[ERROR] {e}"

        pattern = args.get("pattern", "")
        replacement = args.get("replacement", "")
        dry_run = bool(args.get("dry_run", False))
        replace_all = bool(args.get("replace_all", False))
        if not pattern:
            return "[ERROR] 'pattern' is required"
        if replacement is None:
            return "[ERROR] 'replacement' is required"

        try:
            async with file_write_lock(get_workflow_memory_path(self.project_dir, workflow_id)):
                result = update_workflow_memory(
                    self.project_dir,
                    workflow_id,
                    str(pattern),
                    str(replacement),
                    dry_run=dry_run,
                    replace_all=replace_all,
                )
        except Exception as e:
            return f"[ERROR] update_memory failed: {e}"

        payload = {
            "ok": True,
            "workflow_id": workflow_id,
            "path": result["path"],
            "dry_run": result["dry_run"],
            "changed": result["changed"],
            "replacements": result["replacements"],
            "matched_total": result["matched_total"],
            "replace_all": result["replace_all"],
            "chars": result["chars"],
            "over_limit": result["over_limit"],
        }
        if result["over_limit"]:
            payload["warning"] = (
                f"memory exceeds {MEMORY_CHAR_LIMIT} chars after update; "
                "call compress_memory soon"
            )
        return json.dumps(payload, ensure_ascii=False, indent=2)
    
    async def _read_document(self, file_path: Path) -> str:
        """Convert PDF/DOCX/PPTX/XLSX to markdown via markitdown"""
        try:
            return await self._convert_document_with_markitdown(file_path)
        except ImportError:
            return "[Error] markitdown 未安装。请运行: pip install 'markitdown[all]'"
        except asyncio.TimeoutError:
            if file_path.suffix.lower() in ('.xlsx', '.xls'):
                fallback = await self._read_xlsx_fallback(file_path)
                if fallback:
                    return fallback
            return (
                f"[Error] 文档解析超时（>{self._document_read_timeout:g}s）: {file_path.name}\n"
                "suggested_fix=Increase SKILL_TOOLS_DOCUMENT_READ_TIMEOUT, or convert the document with a dedicated parser first."
            )
        except Exception as e:
            # Fallback for xlsx: read directly with openpyxl (data_only, ignore styles)
            if file_path.suffix.lower() in ('.xlsx', '.xls'):
                fallback = await self._read_xlsx_fallback(file_path)
                if fallback:
                    return fallback
            return f"[Error] 文档解析失败: {_format_exception(e)}"

    async def _convert_document_with_markitdown(self, file_path: Path) -> str:
        """Run MarkItDown in a subprocess so slow/corrupt documents cannot block MCP."""
        converter_code = r'''
import sys
import os

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    from markitdown import MarkItDown
except ImportError:
    print("__AGENTCLAW_MARKITDOWN_IMPORT_ERROR__", file=sys.stderr)
    sys.stderr.flush()
    os._exit(42)

try:
    result = MarkItDown().convert(sys.argv[1])
    sys.stdout.write(result.text_content or "")
    sys.stdout.flush()
    sys.stderr.flush()
    os._exit(0)
except Exception as exc:
    print(f"__AGENTCLAW_MARKITDOWN_ERROR__ {type(exc).__name__}: {exc}", file=sys.stderr)
    sys.stderr.flush()
    os._exit(1)
'''
        cmd = [sys.executable, "-c", converter_code, str(file_path)]
        converter_env = dict(os.environ)
        converter_env["PYTHONIOENCODING"] = "utf-8"
        converter_env.setdefault("PYTHONUTF8", "1")
        if sys.platform == "win32":
            proc = await run_subprocess_in_executor(
                cmd,
                timeout=self._document_read_timeout,
                env=converter_env,
            )
            stdout, stderr, returncode = proc.stdout, proc.stderr, proc.returncode
        else:
            try:
                proc = await asyncio.create_subprocess_exec(
                    *cmd,
                    stdin=asyncio.subprocess.DEVNULL,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                    env=converter_env,
                    start_new_session=True,
                )
            except NotImplementedError:
                proc = await run_subprocess_in_executor(
                    cmd,
                    timeout=self._document_read_timeout,
                    env=converter_env,
                )
                stdout, stderr, returncode = proc.stdout, proc.stderr, proc.returncode
            else:
                try:
                    stdout, stderr = await asyncio.wait_for(
                        proc.communicate(),
                        timeout=self._document_read_timeout,
                    )
                except asyncio.TimeoutError:
                    await _terminate_async_process_tree(proc)
                    raise
                returncode = proc.returncode

        stderr_text = _decode_process_output(stderr).strip()
        if returncode == 42:
            raise ImportError(stderr_text)
        if returncode != 0:
            raise RuntimeError(stderr_text or f"markitdown exited with code {returncode}")
        return _decode_process_output(stdout)

    async def _compress_memory(self, args: dict) -> str:
        try:
            workflow_id = self._resolve_memory_workflow_id(args)
        except ValueError as e:
            return f"[ERROR] {e}"

        dry_run = bool(args.get("dry_run", False))
        original = read_workflow_memory(self.project_dir, workflow_id)
        if not original.strip():
            return json.dumps({"ok": True, "workflow_id": workflow_id, "changed": False, "chars": 0}, ensure_ascii=False, indent=2)

        prompt = (
            "You are compressing a workflow memory.md file.\n"
            "Keep critical facts, durable preferences, constraints, identities, file paths, API details, and unfinished tasks.\n"
            "Remove repetition and transient chatter.\n"
            f"Output plain markdown under {MEMORY_COMPRESS_TARGET} characters.\n"
            "Do not lose key information.\n\n"
            f"<memory_md>\n{original}\n</memory_md>"
        )
        manager = self._get_default_llm_manager()
        response = await manager.invoke(messages=[{"role": "user", "content": prompt}])
        compressed = response.content if hasattr(response, "content") else str(response)
        compressed = compressed.strip()

        if len(compressed) >= MEMORY_COMPRESS_TARGET:
            retry_prompt = (
                f"Rewrite the memory below to under {MEMORY_COMPRESS_TARGET} characters. "
                "Keep all key facts. Output markdown only.\n\n"
                f"<memory_md>\n{compressed}\n</memory_md>"
            )
            retry = await manager.invoke(messages=[{"role": "user", "content": retry_prompt}])
            compressed = (retry.content if hasattr(retry, "content") else str(retry)).strip()

        if len(compressed) >= MEMORY_COMPRESS_TARGET:
            return (
                f"[ERROR] compress_memory result is still too long "
                f"({len(compressed)} chars >= {MEMORY_COMPRESS_TARGET})"
            )

        if dry_run:
            result = {
                "path": str(get_workflow_memory_path(self.project_dir, workflow_id)),
                "chars": len(compressed),
                "over_limit": len(compressed) > MEMORY_CHAR_LIMIT,
            }
        else:
            async with file_write_lock(get_workflow_memory_path(self.project_dir, workflow_id)):
                result = write_workflow_memory(self.project_dir, workflow_id, compressed)
        payload = {
            "ok": True,
            "workflow_id": workflow_id,
            "dry_run": dry_run,
            "changed": compressed != original,
            "chars_before": len(original),
            "chars_after": len(compressed),
            "path": result["path"],
            "over_limit": result["over_limit"],
        }
        return json.dumps(payload, ensure_ascii=False, indent=2)

    async def _read_xlsx_fallback(self, file_path: Path) -> Optional[str]:
        """Fallback xlsx reader using openpyxl data_only mode"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"## Sheet: {sheet_name}\n")
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(row)
                if not rows:
                    parts.append("(empty sheet)\n")
                    continue
                # Build markdown table
                header = rows[0]
                col_names = [str(c) if c is not None else "" for c in header]
                parts.append("| " + " | ".join(col_names) + " |")
                parts.append("| " + " | ".join(["---"] * len(col_names)) + " |")
                for row in rows[1:]:
                    cells = [str(c) if c is not None else "" for c in row]
                    # Pad or truncate to match header length
                    while len(cells) < len(col_names):
                        cells.append("")
                    parts.append("| " + " | ".join(cells[:len(col_names)]) + " |")
                parts.append("")
            wb.close()
            return "\n".join(parts)
        except Exception:
            return None

    def _get_vl_client(self):
        """Lazy init VL model client from models.json"""
        if self._vl_client is not None:
            return self._vl_client

        if not self.models_config or not self.models_config.exists():
            return None

        try:
            import json
            config = json.loads(self.models_config.read_text(encoding='utf-8'))
            vision_id = config.get("vision")
            if not vision_id:
                return None

            models = config.get("models", [])
            vl_cfg = next((m for m in models if m.get("id") == vision_id), None)
            if not vl_cfg:
                return None

            from openai import OpenAI
            self._vl_client = {
                "client": OpenAI(api_key=vl_cfg["api_key"], base_url=vl_cfg.get("base_url")),
                "model": vl_cfg["model"],
            }
            return self._vl_client
        except Exception as e:
            logger.warning(f"[skill-tools] VL 模型初始化失败: {e}")
            return None

    async def _read_image(self, file_path: Path, prompt: str) -> str:
        """Analyze image via VL model"""
        vl = self._get_vl_client()
        if not vl:
            return (f"[Image file: {file_path.name}, {file_path.stat().st_size} bytes]\n"
                    f"[Warning] 无法分析图片：未配置视觉模型(vision)。"
                    f"请在 models.json 中添加 \"vision\": \"<model_id>\" 并配置对应的视觉模型。")

        try:
            import base64
            data = base64.b64encode(file_path.read_bytes()).decode('utf-8')
            mime = f"image/{file_path.suffix.lstrip('.').lower()}"
            if mime == "image/jpg":
                mime = "image/jpeg"

            resp = vl["client"].chat.completions.create(
                model=vl["model"],
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{data}"}},
                    ],
                }],
                max_tokens=2048,
            )
            return f"[Image: {file_path.name}]\n{resp.choices[0].message.content}"
        except Exception as e:
            return f"[Error] 图片分析失败: {e}"

    async def _write_file(self, args: dict) -> str:
        """Write file content"""
        path = args.get("path", "")
        content = args.get("content", "")
        allow_large = bool(args.get("allow_large", False))
        overwrite_existing = bool(args.get("overwrite_existing", False))
        
        if not path:
            return "[Error] 'path' is required"
        if not isinstance(path, str):
            return (
                "[Error] write_file validation failed: 'path' must be string\n"
                "error_class=parameter_validation\n"
                "root_cause_hint=path must be a relative file path\n"
                "suggested_fix=Use {'path':'workflows/nl2sql.py','content':'...'}"
            )
        if Path(path).is_absolute():
            return (
                "[Error] write_file validation failed: absolute path is not allowed\n"
                "error_class=parameter_validation\n"
                "root_cause_hint=path must be relative to current working directory\n"
                "suggested_fix=Use project-relative path like 'workflows/nl2sql.py'"
            )
        if "\n" in path or "\r" in path:
            return (
                "[Error] write_file validation failed: path contains newline characters\n"
                "error_class=parameter_validation\n"
                "root_cause_hint=path/content may be swapped\n"
                "suggested_fix=Put file path in 'path' and full source in 'content'"
            )
        if len(path) > 260:
            return (
                "[Error] write_file validation failed: path is too long\n"
                "error_class=parameter_validation\n"
                "root_cause_hint=path/content may be swapped or malformed\n"
                "suggested_fix=Use short relative path; keep source text in 'content'"
            )
        if not isinstance(content, str):
            return (
                "[Error] write_file validation failed: 'content' must be string\n"
                "error_class=parameter_validation\n"
                "root_cause_hint=content must contain text payload\n"
                "suggested_fix=Use {'path':'workflows/nl2sql.py','content':'from agentclaw import ...'}"
            )

        content_bytes = len(content.encode("utf-8"))
        if content_bytes > self._write_file_max_bytes and not allow_large:
            return (
                "[Error] write_file validation failed: content is too large for one-shot write\n"
                "error_class=content_too_large\n"
                f"max_bytes={self._write_file_max_bytes}\n"
                f"actual_bytes={content_bytes}\n"
                "root_cause_hint=large one-shot writes are unstable and harder to recover\n"
                "suggested_fix=Split content into smaller deterministic edits, or retry with allow_large=true only when intentional"
            )
        
        file_path = self.working_dir / path
        try:
            async with file_write_lock(file_path):
                if file_path.exists():
                    if file_path.is_dir():
                        return (
                            "[Error] write_file validation failed: target path is a directory\n"
                            "error_class=path_is_directory\n"
                            "root_cause_hint=write_file expects a file path\n"
                            "suggested_fix=Use a file path under that directory"
                        )
                    try:
                        existing_content = file_path.read_text(encoding='utf-8')
                    except Exception:
                        existing_content = None

                    if existing_content is not None and existing_content == content:
                        return f"[OK] Unchanged: {path} (content identical)"

                    if not overwrite_existing:
                        return (
                            "[Error] write_file validation failed: target file already exists\n"
                            "error_class=existing_file_overwrite_requires_opt_in\n"
                            "root_cause_hint=existing file rewrites are high-risk and often replace too much context\n"
                            "suggested_fix=Use replace_in_file/update_code for local edits, or set overwrite_existing=true for intentional full rewrite"
                        )

                file_path.parent.mkdir(parents=True, exist_ok=True)
                file_path.write_text(content, encoding='utf-8')
            overwrite_warning = ""
            if overwrite_existing:
                overwrite_warning = (
                    "\nwarning=existing_file_overwritten"
                    "\nhint=for follow-up fixes on this file, prefer replace_in_file/update_code instead of another write_file overwrite"
                )
            if content_bytes > self._write_file_max_bytes and allow_large:
                return (
                    f"[OK] Written: {path} ({len(content)} chars, {content_bytes} bytes)\n"
                    "warning=large_write_override_used"
                    f"{overwrite_warning}"
                )
            return f"[OK] Written: {path} ({len(content)} chars, {content_bytes} bytes){overwrite_warning}"
        except Exception as e:
            return f"Write failed: {e}"

    async def _write_code(self, args: dict) -> str:
        """Write code with automatic HTML entity sanitization and pre-write syntax check."""
        path = args.get("path", "")
        content = args.get("content", "")
        overwrite_existing = bool(args.get("overwrite_existing", False))

        # --- parameter validation (reuse write_file's checks) ---
        if not path or not isinstance(path, str):
            return "[Error] 'path' is required and must be a string"
        if Path(path).is_absolute():
            return "[Error] write_code: absolute path not allowed, use relative path"
        if "\n" in path or "\r" in path or len(path) > 260:
            return "[Error] write_code: invalid path (newline or too long)"
        if not isinstance(content, str):
            return "[Error] write_code: 'content' must be a string"

        # --- 1. sanitize HTML entities ---
        sanitized = html.unescape(content)
        was_sanitized = sanitized != content

        # --- 2. detect \\n double-escape (warning only, common model mistake) ---
        double_escape_warning = ""
        if "'\\\\n'" in sanitized or '"\\\\n"' in sanitized:
            double_escape_warning = (
                "\nwarning=double_escaped_newline_detected"
                "\nhint=Found '\\\\n' in code which is literal backslash-n, not a newline. "
                "If you intended actual newlines, use '\\n' instead."
            )

        # --- 3. pre-write syntax check (Python only for now) ---
        suffix = Path(path).suffix.lower()
        if suffix == ".py":
            try:
                compile(sanitized, path, "exec")
            except SyntaxError as e:
                # Build readable error context
                lines = sanitized.splitlines()
                err_line = e.lineno or 0
                context_parts = []
                start = max(1, err_line - 3)
                end = min(len(lines), err_line + 3)
                for ln in range(start, end + 1):
                    prefix = ">  " if ln == err_line else "   "
                    text = lines[ln - 1] if 0 <= (ln - 1) < len(lines) else ""
                    context_parts.append(f"{prefix}{ln:>4} | {text}")
                context_str = "\n".join(context_parts)

                # Smart hint
                hint = ""
                if err_line > 0 and 0 <= (err_line - 1) < len(lines):
                    err_text = lines[err_line - 1]
                    if "&quot;" in err_text or "&#x27;" in err_text or "&gt;" in err_text:
                        hint = "\nHINT: Line contains HTML entities. Sanitization may not have fixed all issues."
                    elif "\\\\n" in err_text or "\\\\t" in err_text:
                        hint = "\nHINT: Line contains double-escaped sequences (\\\\n). Use single \\n for newlines."

                return (
                    f"[SYNTAX_ERROR] Pre-write check failed — file NOT written\n"
                    f"File: {path}:{err_line}:{e.offset or 0}\n"
                    f"Error: {e.msg}\n\n"
                    f"{context_str}"
                    f"{hint}"
                    f"{double_escape_warning}"
                )
        elif suffix == ".json":
            try:
                import json as json_mod
                json_mod.loads(sanitized)
            except json_mod.JSONDecodeError as e:
                return (
                    f"[SYNTAX_ERROR] Pre-write JSON check failed — file NOT written\n"
                    f"File: {path}:{e.lineno}:{e.colno}\n"
                    f"Error: {e.msg}"
                )

        # --- 4. write file ---
        file_path = self.working_dir / path
        try:
            async with file_write_lock(file_path):
                if file_path.exists():
                    if file_path.is_dir():
                        return "[Error] write_code: target path is a directory"
                    if not overwrite_existing:
                        return (
                            "[Error] write_code: file exists, set overwrite_existing=true "
                            "or use replace_in_file/update_code for local edits"
                        )

                file_path.parent.mkdir(parents=True, exist_ok=True)
                file_path.write_text(sanitized, encoding="utf-8")

            content_bytes = len(sanitized.encode("utf-8"))
            sanitize_note = ""
            if was_sanitized:
                sanitize_note = "\nnote=html_entities_auto_fixed (content was sanitized before writing)"

            # 对 Python/JSON 文件，标注语法已通过（避免 LLM 再调 syntax_check）
            syntax_note = ""
            if suffix == ".py":
                syntax_note = "\nsyntax_check: PASSED (pre-write validation — no separate syntax_check needed)"
            elif suffix == ".json":
                syntax_note = "\njson_check: PASSED (pre-write validation)"

            return (
                f"[OK] Written: {path} ({len(sanitized)} chars, {content_bytes} bytes)"
                f"{syntax_note}"
                f"{sanitize_note}"
                f"{double_escape_warning}"
            )
        except Exception as e:
            return f"[Error] write_code failed: {e}"

    async def _list_files(self, args: dict) -> str:
        """List directory contents"""
        path = args.get("path", ".")
        skill_name = (args.get("skill_name") or "").strip()

        resolved_path = None
        candidates = []

        if skill_name:
            # skill_name 给出时：skill 目录优先 → fallback working_dir
            skill_obj, _ = self._get_skill_binding(skill_name)
            if skill_obj and skill_obj.path and skill_obj.path.exists():
                skill_candidate = (skill_obj.path / path).resolve()
                if skill_candidate.exists():
                    resolved_path = skill_candidate
                candidates.append(skill_candidate)

            # fallback 到 working_dir
            if not resolved_path:
                fallback, fallback_candidates = self._resolve_existing_path(path)
                candidates.extend(fallback_candidates)
                if fallback:
                    resolved_path = fallback
        else:
            resolved_path, candidates = self._resolve_existing_path(path)

        if not resolved_path:
            payload = {
                "ok": False,
                "error_class": "path_not_found",
                "requested_path": path,
                "working_dir": str(self.working_dir),
                "root_cause_hint": "directory lookup failed under working_dir and ancestor directories",
                "suggested_fix": "Use an existing directory path or inspect current root with list_files(path='.')",
            }
            return "Directory not found\n" + json.dumps(payload, ensure_ascii=False, indent=2)

        if resolved_path.is_file():
            size = resolved_path.stat().st_size
            return (
                f"📄 {resolved_path.name} ({size} bytes)\n"
                f"[hint] '{path}' resolves to a file, not a directory.\n"
                f"[hint] Use read_file(path='{path}') to view content.\n"
                f"[resolved_path] {resolved_path}"
            )

        # uploads 目录：调 API 获取原始文件名
        if self._is_uploads_dir(resolved_path):
            api_listing = await self._list_uploads_via_api()
            if api_listing is not None:
                return api_listing

        try:
            items = []
            is_root = resolved_path == self.working_dir
            for item in sorted(resolved_path.iterdir()):
                if item.is_dir():
                    name = item.name
                    # 根目录下 .storage 加注释说明
                    if is_root and name in (".storage", ".storage/"):
                        items.append(f"📁 {name}/  # 文件存储目录（上传文件、知识库文档等），由平台自动管理")
                    else:
                        items.append(f"📁 {name}/")
                else:
                    size = item.stat().st_size
                    items.append(f"📄 {item.name} ({size} bytes)")
            listing = "\n".join(items) if items else "(empty directory)"
            direct_path = self.working_dir / path
            try:
                direct_resolved = direct_path.resolve(strict=False)
            except Exception:
                direct_resolved = direct_path
            if str(direct_resolved) != str(resolved_path):
                return (
                    f"[path_resolved] requested={path} resolved={resolved_path}\n"
                    f"{listing}"
                )
            return listing
        except Exception as e:
            return f"List failed: {e}"

    def _is_uploads_dir(self, resolved_path: Path) -> bool:
        """检测是否为 .storage/uploads 目录"""
        if resolved_path.name != "uploads":
            return False
        parent = resolved_path.parent
        return parent.name.startswith(".storage")

    async def _list_uploads_via_api(self) -> Optional[str]:
        """调用 /api/upload/list 获取上传文件列表（含原始文件名）"""
        import urllib.request
        import urllib.error

        base_url = self._resolve_server_base_url(internal=True)
        url = f"{base_url}/_internal/api/upload/list"
        try:
            req = urllib.request.Request(url, method="GET")
            req.add_header("Content-Type", "application/json")
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            files = data.get("files", [])
            if not files:
                return "(empty directory)"
            items = []
            for f in files:
                storage_name = f.get("file_path", "").split("/")[-1]
                original = f.get("original_name", storage_name)
                size = f.get("size", 0)
                mime = f.get("mime_type", "")
                file_id = f.get("id", "")
                display = f"📄 {storage_name} → {original} ({size} bytes)"
                if mime:
                    display += f" [{mime}]"
                if file_id:
                    display += f" (id: {file_id})"
                items.append(display)
            return "\n".join(items)
        except Exception as e:
            logger.debug(f"调用 upload/list API 失败，回退到文件系统列表: {e}")
            return None

    async def _read_skill_file(self, args: dict) -> str:
        """Read a skill's documentation file"""
        skill_name = args.get("skill_name")
        file_name = args.get("file_name", "SKILL.md")

        if not skill_name:
            return "[ERROR] skill_name is required"

        def _append_search_paths(source: str, skill_root: Path):
            requested = Path(file_name)
            candidates = [skill_root / requested]
            if len(requested.parts) == 1 and file_name != "SKILL.md":
                candidates.append(skill_root / "references" / file_name)
            for candidate in candidates:
                if not any(existing == candidate for _, existing in search_paths):
                    search_paths.append((source, candidate))

        # Search locations
        search_paths = []

        # 1. Project skills directory
        if self.skills_dir and self.skills_dir.exists():
            # 获取实际的目录名（可能与 skill_name 不同）
            skill_dir_name = skill_name
            if self._skill_manager:
                actual_dir = self._skill_manager.get_skill_dir_name(skill_name)
                if actual_dir:
                    skill_dir_name = actual_dir
            project_skill_root = self.skills_dir / skill_dir_name
            _append_search_paths("project", project_skill_root)

        # 2. Builtin skills directory
        try:
            import agentclaw
            agentclaw_root = Path(agentclaw.__file__).parent
            builtin_skill_root = agentclaw_root / "skills" / "builtin_skills" / skill_name
            _append_search_paths("builtin", builtin_skill_root)
        except Exception:
            pass

        # Try each location
        for source, file_path in search_paths:
            if file_path.exists() and file_path.is_file():
                try:
                    content = file_path.read_text(encoding="utf-8")
                    content = self._substitute_skill_runtime_variables(content, skill_name)
                    numbered = _add_line_numbers(content)
                    return f"[OK] Read skill file: {skill_name}/{file_name} (source: {source})\n\n{numbered}"
                except Exception as e:
                    return f"[ERROR] Failed to read {file_path}: {e}"

        # Not found
        searched = "\n".join(f"  - {source}: {path}" for source, path in search_paths)
        return f"[ERROR] Skill file not found: {skill_name}/{file_name}\n\nSearched locations:\n{searched}"

    async def _execute_sudo_command(self, args: dict) -> str:
        """Execute a command with sudo privileges using credential from Harness confirmation."""
        command = (args.get("command") or "").strip()
        if not command:
            return "[ERROR] Missing 'command' parameter"

        timeout = args.get("timeout")
        if timeout is None:
            timeout = os.getenv("SUDO_COMMAND_TIMEOUT", "60")
        try:
            timeout = float(timeout)
            if timeout <= 0:
                timeout = 60.0
        except (TypeError, ValueError):
            timeout = 60.0

        # 从临时文件读取 sudo 密码（由 Harness 工具确认流程写入）
        import tempfile
        sudo_file = Path(tempfile.gettempdir()) / ".agentclaw_sudo"
        if not sudo_file.exists():
            return (
                "[ERROR] No sudo password available. "
                "Enable Harness tool confirmation and approve this sudo tool call first."
            )

        try:
            sudo_password = sudo_file.read_text(encoding="utf-8")
        except Exception as e:
            return f"[ERROR] Failed to read sudo credential: {e}"
        finally:
            # 用完即删
            try:
                sudo_file.unlink(missing_ok=True)
            except Exception:
                pass

        if not sudo_password:
            return "[ERROR] Sudo credential file is empty. Approve this sudo tool call again through Harness confirmation."

        logger.info(f"execute_sudo_command: executing sudo command: {command}")

        try:
            full_command = f"echo '{sudo_password}' | sudo -S {command}"
            proc = await asyncio.create_subprocess_shell(
                full_command,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)

            stdout_text = stdout.decode("utf-8", errors="replace").strip()
            stderr_text = stderr.decode("utf-8", errors="replace").strip()

            # 过滤 sudo 密码提示信息
            if stderr_text:
                lines = stderr_text.split("\n")
                filtered = [
                    line for line in lines
                    if not line.startswith("[sudo]") and "password for" not in line.lower()
                ]
                stderr_text = "\n".join(filtered).strip()

            if proc.returncode == 0:
                return stdout_text if stdout_text else "[Command executed successfully, no output]"
            else:
                error_msg = stderr_text if stderr_text else f"Command failed with exit code {proc.returncode}"
                return f"[ERROR] {error_msg}"

        except asyncio.TimeoutError:
            return f"[ERROR] Command execution timeout ({int(timeout)}s)"
        except Exception as e:
            return f"[ERROR] Command execution failed: {e}"

    async def _create_download_url(self, args: dict) -> str:
        """Generate a temporary download URL for a local file."""
        import base64
        import mimetypes
        import uuid

        file_path_str = args.get("path", "")
        if not file_path_str:
            return "[ERROR] 'path' is required"

        file_path = Path(file_path_str)
        if not file_path.is_absolute():
            file_path = self.working_dir / file_path

        # Security check: prevent path traversal
        try:
            resolved = file_path.resolve()
            working_resolved = self.working_dir.resolve()
            try:
                resolved.relative_to(working_resolved)
            except ValueError:
                return f"[ERROR] Access denied: file must be within working directory ({working_resolved})"
        except Exception:
            return "[ERROR] Invalid file path"

        if not file_path.exists():
            return f"[ERROR] File not found: {file_path}"
        if not file_path.is_file():
            return f"[ERROR] Not a file: {file_path}"

        size_error = validate_download_file_size(file_path)
        if size_error:
            return size_error

        try:
            content = file_path.read_bytes()
        except Exception as e:
            return f"[ERROR] Failed to read file: {e}"

        filename = args.get("filename") or file_path.name
        ttl = normalize_download_ttl(args.get("ttl", 3600))
        content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
        token = uuid.uuid4().hex

        try:
            redis_client = DatabaseManager.create_sync_redis_client(RedisConfig.from_env())
            key = f"download:{token}"
            redis_client.hset(key, mapping={
                "content": base64.b64encode(content).decode("ascii"),
                "filename": filename,
                "content_type": content_type,
            })
            redis_client.expire(key, ttl)
        except Exception as e:
            return f"[ERROR] Redis error: {e}"

        base_url = os.getenv("DOWNLOAD_BASE_URL", "/api/download")
        url = f"{base_url}/{token}"

        return (
            f"Download URL created:\n"
            f"  URL: {url}\n"
            f"  File: {filename}\n"
            f"  Type: {content_type}\n"
            f"  Size: {len(content)} bytes\n"
            f"  Expires in: {ttl}s"
        )

    async def run(self):
        """Run the MCP server"""
        logger.info(f"[skill-tools] Starting MCP server (stdio)")
        logger.info(f"[skill-tools] Skills dir: {self.skills_dir}")
        logger.info(f"[skill-tools] Working dir: {self.working_dir}")

        async with stdio_server() as (read_stream, write_stream):
            # Start skill loading in background (non-blocking)
            asyncio.create_task(self._init_skill_manager())

            await self._server.run(
                read_stream,
                write_stream,
                self._server.create_initialization_options()
            )
